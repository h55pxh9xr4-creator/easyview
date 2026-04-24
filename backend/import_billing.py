"""Easy View 빌링현황.xlsm → billing DB 일괄 import.

- 데이터 소스: Asset/Easy View 빌링현황.xlsm (git에 커밋 안 함, 로컬 전용)
- 재실행 안전: (관리번호, report 기준월) 기준 upsert
- 시트:
  * Billing List → billing_entry (status="대기중")
  * 완료리스트 → billing_entry (status="완료")
  * Master → billing_master
  * 특이사항 법인 → billing_exception
- 실행: python import_billing.py (backend 디렉터리에서)
"""
import sys
import warnings
from datetime import datetime, date
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning)

from database import SessionLocal, engine, Base
from billing_models import BillingEntry, BillingMaster, BillingException

BACKEND_DIR = Path(__file__).resolve().parent
XLSM_PATH = BACKEND_DIR.parent / "Asset" / "Easy View 빌링현황.xlsm"


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _to_str(v):
    if v is None:
        return None
    return str(v).strip()


def _to_amount(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def import_billing():
    if not XLSM_PATH.exists():
        print(f"[빌링] 엑셀 파일 없음: {XLSM_PATH}")
        print("       Asset/ 폴더에 'Easy View 빌링현황.xlsm' 을 두고 다시 실행하세요.")
        return
    import openpyxl
    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=False)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    ins_entry = upd_entry = 0
    ins_master = upd_master = 0
    ins_exc = 0

    try:
        # ── Billing List (대기중) ─────────────────────────────────
        if "Billing List" in wb.sheetnames:
            ws = wb["Billing List"]
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i == 1:
                    continue
                if not row or not row[0] or not row[2]:
                    continue
                parent = _to_str(row[0])
                subsidiary = _to_str(row[1])
                mgmt_no = _to_str(row[2])
                assignee = _to_str(row[3])
                report_ym = _to_str(row[4])
                delivery_ym = _to_str(row[5])
                billing_date = _to_date(row[6])
                invoice_date = _to_date(row[7])
                status = _to_str(row[8]) or "빌링대기중"
                deposit_date = _to_date(row[9])
                memo = _to_str(row[10])
                amount = _to_amount(row[11])
                invoice_request_day = _to_str(row[12])
                invoice_manager = _to_str(row[13])
                manager_email = _to_str(row[14])
                manager_phone = _to_str(row[15])

                # pending 엔트리는 transfer_at=NULL 기준으로 dedup
                exists = db.query(BillingEntry).filter(
                    BillingEntry.mgmt_no == mgmt_no,
                    BillingEntry.report_ym == report_ym,
                    BillingEntry.transfer_at.is_(None),
                    BillingEntry.is_completed == False,
                ).first()
                payload = dict(
                    parent=parent, subsidiary=subsidiary, mgmt_no=mgmt_no,
                    assignee=assignee, report_ym=report_ym, delivery_ym=delivery_ym,
                    billing_date=billing_date, invoice_date=invoice_date,
                    status=status, deposit_date=deposit_date,
                    amount=amount, invoice_request_day=invoice_request_day,
                    invoice_manager=invoice_manager, manager_email=manager_email,
                    manager_phone=manager_phone, is_completed=False,
                )
                if exists:
                    # 🔒 UI 편집 보호: memo/deposit_date는 엑셀이 비었으면 기존 값 유지
                    if memo:
                        exists.memo = memo
                    if deposit_date:
                        exists.deposit_date = deposit_date
                    for k, v in payload.items():
                        setattr(exists, k, v)
                    upd_entry += 1
                else:
                    payload["memo"] = memo
                    db.add(BillingEntry(**payload))
                    ins_entry += 1
                db.flush()  # 같은 run 안에서 즉시 가시화

        # ── 완료리스트 ────────────────────────────────────────────
        if "완료리스트" in wb.sheetnames:
            ws = wb["완료리스트"]
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i == 1:
                    continue
                if not row or not row[1] or not row[3]:
                    continue
                transfer_at = _to_date(row[0])
                parent = _to_str(row[1])
                subsidiary = _to_str(row[2])
                mgmt_no = _to_str(row[3])
                assignee = _to_str(row[4])
                report_ym = _to_str(row[5])
                delivery_ym = _to_str(row[6])
                billing_date = _to_date(row[7])
                invoice_date = _to_date(row[8])
                status = _to_str(row[9]) or "완료"
                deposit_date = _to_date(row[10])
                memo = _to_str(row[11])
                amount = _to_amount(row[12])
                invoice_request_day = _to_str(row[13])
                invoice_manager = _to_str(row[14])

                # completed 엔트리는 (mgmt_no, report_ym, transfer_at) 조합으로 dedup
                # → 같은 고객의 같은 기준월이라도 이관일이 다르면 별개 이력으로 보관
                exists = db.query(BillingEntry).filter(
                    BillingEntry.mgmt_no == mgmt_no,
                    BillingEntry.report_ym == report_ym,
                    BillingEntry.transfer_at == transfer_at,
                    BillingEntry.is_completed == True,
                ).first()
                payload = dict(
                    parent=parent, subsidiary=subsidiary, mgmt_no=mgmt_no,
                    assignee=assignee, report_ym=report_ym, delivery_ym=delivery_ym,
                    billing_date=billing_date, invoice_date=invoice_date,
                    status=status, deposit_date=deposit_date,
                    amount=amount, invoice_request_day=invoice_request_day,
                    invoice_manager=invoice_manager, is_completed=True,
                    transfer_at=transfer_at,
                )
                if exists:
                    if memo:
                        exists.memo = memo
                    if deposit_date:
                        exists.deposit_date = deposit_date
                    for k, v in payload.items():
                        setattr(exists, k, v)
                    upd_entry += 1
                else:
                    payload["memo"] = memo
                    db.add(BillingEntry(**payload))
                    ins_entry += 1
                db.flush()

        # ── Master ───────────────────────────────────────────────
        if "Master" in wb.sheetnames:
            ws = wb["Master"]
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i == 1:
                    continue
                if not row or not row[0]:
                    continue
                mgmt_no = _to_str(row[0])
                company = _to_str(row[1])
                parent = _to_str(row[2])
                amount = _to_amount(row[3])
                invoice_request_day = _to_str(row[4])
                invoice_manager = _to_str(row[5])
                manager_email = _to_str(row[6])
                manager_phone = _to_str(row[7])

                exists = db.query(BillingMaster).filter(BillingMaster.mgmt_no == mgmt_no).first()
                payload = dict(
                    mgmt_no=mgmt_no, company=company, parent=parent,
                    amount=amount, invoice_request_day=invoice_request_day,
                    invoice_manager=invoice_manager, manager_email=manager_email,
                    manager_phone=manager_phone,
                )
                if exists:
                    for k, v in payload.items():
                        setattr(exists, k, v)
                    upd_master += 1
                else:
                    db.add(BillingMaster(**payload))
                    ins_master += 1

        # ── 특이사항 법인 ────────────────────────────────────────
        if "특이사항 법인" in wb.sheetnames:
            ws = wb["특이사항 법인"]
            # 특이사항은 재실행 시 전체 교체 (key가 없어서 dedup 어려움)
            db.query(BillingException).delete()
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i == 1:
                    continue
                if not row or not row[0] or not row[2]:
                    continue
                no = int(row[0]) if isinstance(row[0], (int, float)) else None
                category = _to_str(row[1])
                parent = _to_str(row[2])
                note = _to_str(row[3])
                db.add(BillingException(
                    no=no, category=category, parent=parent, note=note,
                ))
                ins_exc += 1

        db.commit()
    finally:
        db.close()

    return {
        "entry_inserted": ins_entry, "entry_updated": upd_entry,
        "master_inserted": ins_master, "master_updated": upd_master,
        "exception_inserted": ins_exc,
    }


if __name__ == "__main__":
    result = import_billing()
    if result:
        print(f"[빌링 import 완료] {result}")
